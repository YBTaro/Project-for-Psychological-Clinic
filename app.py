import tkinter as tk
from tkinter import filedialog, messagebox
import pandas as pd
from openpyxl import Workbook
import os

# 建立主視窗
root = tk.Tk()
root.title("Excel 處理工具")
root.geometry("400x200")

# 上傳檔案
file_path = ""
def upload_file():
    global file_path
    file_path = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx")])
    if file_path:
        messagebox.showinfo("檔案已上傳", f"檔案路徑: {file_path}")

def process_file():
    if not file_path:
        messagebox.showerror("錯誤", "請先上傳檔案")
        return

    try:
        # 讀取 Excel 檔案
        df = pd.read_excel(file_path)
        # cleaning data
        df[["計畫自費(同一人多筆算一位)","健保人數（含壓單、舌下健保初診）","舌下錠自費人數","初診人數(新病歷號)"]] = df[["計畫自費(同一人多筆算一位)","健保人數（含壓單、舌下健保初診）","舌下錠自費人數","初診人數(新病歷號)"]].fillna(0)
        df['計畫自費(同一人多筆算一位)'] = df['計畫自費(同一人多筆算一位)'].astype(int)
        df["計算後的當天人數"] = df["健保人數（含壓單、舌下健保初診）"]+df["舌下錠自費人數"]+0.5*df["計畫自費(同一人多筆算一位)"]
        df = df[df['醫師'] != "休診"]
        # 重要列表
        week_list = df['第幾周'].unique().tolist()
        doctor_list = ['Dr. 張', 'Dr. 李', 'Dr. 陳', 'Dr. 林', 'Dr. 黃', 'Dr. 徐']

        # 建立新的試算表
        wb = Workbook()
        ws = wb.active
        ws.title = "示範工作表"
        index_newFile = 0
        dic_col = {0:"A",
            1:"B",
            2:"C",
            3:"D",
            4:"E",
            5:"F",
            6:"G",
            7:"H"}

        # 開始計算
        for week in week_list:
            df_selected = df[df['第幾周'] == week]

            # 建立新的欄位
            week_cell = "B"+str(1+4*index_newFile)+":G"+str(1+4*index_newFile)
            ws.merge_cells(week_cell)
            ws["B"+str(1+4*index_newFile)] = "第"+str(week_list[index_newFile])+"週每診平均人數"
            for i in range(len(doctor_list)):
                doctor_df = df_selected[df_selected["醫師"] == doctor_list[i]]
                ws[dic_col[i+1]+str(2+4*index_newFile)] = doctor_list[i]
                ws[dic_col[i+1]+str(3+4*index_newFile)] = doctor_df["計算後的當天人數"].mean()
            index_newFile+=1

        # 月平均
        month_cell = "B"+str(1+4*index_newFile)+":H"+str(1+4*index_newFile)
        ws.merge_cells(month_cell)
        ws["B"+str(1+4*index_newFile)] = "本月統計資料"
        ws["A"+str(3+4*index_newFile)] = "本月總診數"
        ws["A"+str(4+4*index_newFile)] = "本月總人數"
        ws["A"+str(5+4*index_newFile)] = "本月每診平均人數"
        for i in range(len(doctor_list)):
            doctor_df = df[df["醫師"] == doctor_list[i]]
            ws[dic_col[i+1]+str(2+4*index_newFile)] = doctor_list[i]
            ws[dic_col[i+1]+str(3+4*index_newFile)] = doctor_df["計算後的當天人數"].count()
            ws[dic_col[i+1]+str(4+4*index_newFile)] = doctor_df["計算後的當天人數"].sum()
            ws[dic_col[i+1]+str(5+4*index_newFile)] = doctor_df["計算後的當天人數"].mean()


        ws[dic_col[len(doctor_list)+1]+str(2+4*index_newFile)] = "全部"
        ws[dic_col[len(doctor_list)+1]+str(3+4*index_newFile)] = df["計算後的當天人數"].count()
        ws[dic_col[len(doctor_list)+1]+str(4+4*index_newFile)] = df["計算後的當天人數"].sum()
        ws[dic_col[len(doctor_list)+1]+str(5+4*index_newFile)] = df["計算後的當天人數"].mean()


        wb_name = os.path.basename(file_path)
        wb_name = os.path.splitext(wb_name)[0]
        wb_name = wb_name + "-報告.xlsx"
        wb.save(wb_name)
        messagebox.showinfo("已完成處理", f"檔案名稱: {wb_name}")

    except Exception as e:
        messagebox.showerror("錯誤", f"處理檔案時發生錯誤: {e}")

# 介面按鈕
upload_btn = tk.Button(root, text="上傳 Excel 檔案", command=upload_file)
process_btn = tk.Button(root, text="執行處理邏輯", command=process_file)

upload_btn.pack(pady=20)
process_btn.pack(pady=20)

# 啟動主迴圈
root.mainloop()